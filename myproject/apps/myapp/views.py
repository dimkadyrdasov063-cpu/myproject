from django.contrib import messages
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Count
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .models import Article, Category, Comment, Favorite
from .forms import ArticleForm, SignUpForm


# ==================== ОСНОВНЫЕ ВЬЮХИ ====================

def article_list(request):
    query = request.GET.get("q")
    category_slug = request.GET.get("category")
    sort = request.GET.get("sort", "new")

    articles = Article.objects.select_related("author", "category").filter(is_published=True)

    if query:
        articles = articles.filter(title__icontains=query)
    if category_slug:
        articles = articles.filter(category__slug=category_slug)

    if sort == "popular":
        articles = articles.annotate(comments_count=Count('comments')).order_by('-views_count', '-comments_count')
    else:
        articles = articles.order_by('-created_at')

    categories = Category.objects.all()

    paginator = Paginator(articles, 10)
    page_obj = paginator.get_page(request.GET.get("page"))

    return render(request, "articles/article_list.html", {
        "articles": page_obj,
        "query": query,
        "categories": categories,
        "current_category": category_slug,
        "current_sort": sort,
    })


def article_detail(request, pk):
    article = get_object_or_404(
        Article.objects.select_related("author", "category").filter(is_published=True),
        pk=pk
    )

    article.views_count += 1
    article.save(update_fields=["views_count"])

    return render(request, "articles/article_detail.html", {"article": article})


@login_required(login_url="articles:login")
def create_article(request):
    if request.method == "POST":
        form = ArticleForm(request.POST)
        if form.is_valid():
            article = form.save(commit=False)
            article.author = request.user
            article.save()
            messages.success(request, "Статья успешно опубликована!")
            return redirect("articles:article_detail", pk=article.pk)
    else:
        form = ArticleForm()

    return render(request, "articles/article_create.html", {"form": form})


def home(request):
    return render(request, "articles/home.html")


def signup(request):
    if request.method == "POST":
        form = SignUpForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect("articles:article_list")
    else:
        form = SignUpForm()

    return render(request, "articles/signup.html", {"form": form})


@login_required
def add_comment(request, pk):
    article = get_object_or_404(Article, pk=pk)
    if request.method == "POST":
        text = request.POST.get("text")
        if text.strip():
            Comment.objects.create(article=article, author=request.user, text=text)
            messages.success(request, "Комментарий добавлен!")
    return redirect("articles:article_detail", pk=pk)


@login_required
def toggle_favorite(request, pk):
    article = get_object_or_404(Article, pk=pk)
    favorite, created = Favorite.objects.get_or_create(user=request.user, article=article)

    if not created:
        favorite.delete()
        messages.info(request, "Удалено из избранного")
    else:
        messages.success(request, "Добавлено в избранное ❤️")

    return redirect("articles:article_detail", pk=pk)


@login_required(login_url="articles:login")
def export_articles_to_excel(request):
    articles = Article.objects.select_related("author", "category").all()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Статьи"

    headers = ["ID", "Заголовок", "Автор", "Категория", "Дата создания", "Дата редактирования", "Опубликована", "Просмотры"]
    worksheet.append(headers)

    for article in articles:
        category_name = article.category.name if article.category else "Без категории"
        row = [
            article.pk,
            article.title,
            article.author.username,
            category_name,
            article.created_at.strftime("%d.%m.%Y %H:%M"),
            article.updated_at.strftime("%d.%m.%Y %H:%M"),
            "Да" if article.is_published else "Нет",
            article.views_count
        ]
        worksheet.append(row)

    widths = [8, 34, 20, 22, 18, 18, 12, 10]
    for column_number, width in enumerate(widths, start=1):
        column_letter = get_column_letter(column_number)
        worksheet.column_dimensions[column_letter].width = width

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="articles.xlsx"'
    workbook.save(response)
    return response


# ==================== ВЬЮХИ КОТОРЫХ НЕ ХВАТАЕТ (добавь если нужно) ====================

@login_required
def article_edit(request, pk):
    article = get_object_or_404(Article, pk=pk, author=request.user)
    if request.method == "POST":
        form = ArticleForm(request.POST, instance=article)
        if form.is_valid():
            form.save()
            messages.success(request, "Статья обновлена!")
            return redirect("articles:article_detail", pk=pk)
    else:
        form = ArticleForm(instance=article)
    return render(request, "articles/article_edit.html", {"form": form, "article": article})


@login_required
def article_delete(request, pk):
    article = get_object_or_404(Article, pk=pk, author=request.user)
    if request.method == "POST":
        article.delete()
        messages.success(request, "Статья удалена!")
        return redirect("articles:article_list")
    return render(request, "articles/article_confirm_delete.html", {"article": article})


@login_required
def my_articles(request):
    articles = Article.objects.filter(author=request.user).order_by('-created_at')
    paginator = Paginator(articles, 10)
    page_obj = paginator.get_page(request.GET.get("page"))
    return render(request, "articles/my_articles.html", {"articles": page_obj})


@login_required
def profile(request):
    favorites = Favorite.objects.filter(user=request.user).select_related('article')
    return render(request, "articles/profile.html", {"favorites": favorites})